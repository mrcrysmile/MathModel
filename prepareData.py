# %%
import numpy as np
import openpyxl as oxl

# 目标数据列
tar_cols = []

# 获取目标数据
def load_tar_xml(path):
    # 操作变量列
    op_cols = []
    op_data = []
    # 性质列
    pro_cols = []
    pro_data = []

    wb = oxl.load_workbook(path)
    tar_sheet = wb.get_sheet_by_name("Sheet1")

    # 获取性质列
    for i in range(3, 17):
        pro_cols.append(tar_sheet.cell(row=4, column=i).value)
    
    # 获取操作变量列
    

# 获取原始数据
def load_ori_xml(path):
    # 操作变量列
    op_cols = []
    op_data = []
    # 性质列
    pro_cols = []
    pro_data = []

    wb = oxl.load_workbook(path)
    # names = wb.get_sheet_names()
    # 操作变量sheet
    op_sheet = wb.get_sheet_by_name("操作变量")
    # 原料sheet
    pro_sheet1 = wb.get_sheet_by_name("原料")
    # 产品sheet
    pro_sheet2 = wb.get_sheet_by_name("产品")
    # 待生吸附剂sheet
    pro_sheet3 = wb.get_sheet_by_name("待生吸附剂")
    # 再生吸附剂sheet
    pro_sheet4 = wb.get_sheet_by_name("再生吸附剂")

    # 处理操作变量sheet
    row = op_sheet.max_row
    col = op_sheet.max_column
    # 读取操作变量列(英文,中文就不处理了)
    for i in range(1, col + 1):
        op_cols.append(op_sheet.cell(row=3, column=i).value)
    # 读取数据
    for i in range(4, row + 1):
        temp_row = []
        for j in range(1, col + 1):
            temp_row.append(op_sheet.cell(row=i, column=j).value)
        op_data.append(temp_row)
    
    # 处理性质sheet
    t1 = []
    t2 = []
    # 原料sheet
    col1 = pro_sheet1.max_column
    for i in range(1, col1):
        pro_cols.append(pro_sheet1.cell(row=1, column=i).value)
        t1.append(pro_sheet1.cell(row=2, column=i).value)
        t2.append(pro_sheet1.cell(row=3, column=i).value)
    # 产品sheet
    col2 = pro_sheet2.max_column
    for i in range(1, col2):
        pro_cols.append(pro_sheet2.cell(row=2, column=i).value)
        t1.append(pro_sheet2.cell(row=3, column=i).value)
        t2.append(pro_sheet2.cell(row=4, column=i).value)
    # 待生吸附剂sheet
    col3 = pro_sheet3.max_column
    for i in range(1, col3):
        pro_cols.append(pro_sheet3.cell(row=2, column=i).value)
        t1.append(pro_sheet3.cell(row=3, column=i).value)
        t2.append(pro_sheet3.cell(row=4, column=i).value)
    # 再生吸附剂sheet
    col4 = pro_sheet4.max_column
    for i in range(1, col4):
        pro_cols.append(pro_sheet4.cell(row=2, column=i).value)
        t1.append(pro_sheet4.cell(row=3, column=i).value)
        t2.append(pro_sheet4.cell(row=4, column=i).value)
    pro_data.append(t1)
    pro_data.append(t2)

    return op_cols, op_data, pro_cols, pro_data


# %%
if __name__ == "__main__":
    pass